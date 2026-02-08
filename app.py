from flask import Flask, request, jsonify, render_template, session, redirect, url_for, flash
from database import Database
from whatsapp_handler import WhatsAppHandler
from config import Config
import os
import json
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config.from_object(Config)

# Initialize
db = Database()
whatsapp = WhatsAppHandler()

# Ensure upload folder exists
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)

def get_next_7_days():
    """Get next 7 days for booking"""
    today = datetime.now()
    dates = []
    for i in range(7):
        date = today + timedelta(days=i)
        label = "Today" if i == 0 else ("Tomorrow" if i == 1 else date.strftime("%d %b, %a"))
        dates.append({
            "value": date.strftime("%Y-%m-%d"),
            "label": label
        })
    return dates

def format_service_list():
    """Format service list for display"""
    text = "*📋 Our Services:*\n\n"
    for key, service in Config.SERVICES.items():
        text += f"{key}. {service['name']}\n"
        text += f"   💰 ₹{service['price']} | ⏱️ {service['duration']}\n\n"
    return text

def get_available_slots(date):
    """Get available time slots for a date"""
    booked_slots = db.get_booked_slots(date)
    available = [slot for slot in Config.TIME_SLOTS if slot not in booked_slots]
    return available

def is_valid_service_input(message):
    """Validate service selection format"""
    try:
        service_nums = [s.strip() for s in message.replace(' ', '').split(',')]
        for num in service_nums:
            if num not in Config.SERVICES:
                return False
        return len(service_nums) > 0
    except:
        return False

# =================== WEBHOOK VERIFICATION ===================

@app.route('/webhook', methods=['GET'])
def verify_webhook():
    """Verify webhook for WhatsApp"""
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    
    if mode == 'subscribe' and token == Config.VERIFY_TOKEN:
        return challenge, 200
    return 'Forbidden', 403

# =================== WEBHOOK HANDLER ===================

@app.route('/webhook', methods=['POST'])
def webhook():
    """Handle incoming WhatsApp messages"""
    try:
        data = request.get_json()
        
        if not data.get('entry'):
            return jsonify({'status': 'ok'}), 200
        
        entry = data['entry'][0]
        changes = entry.get('changes', [])
        
        if not changes:
            return jsonify({'status': 'ok'}), 200
        
        change = changes[0]
        value = change.get('value', {})
        messages = value.get('messages', [])
        
        if not messages:
            return jsonify({'status': 'ok'}), 200
        
        message = messages[0]
        from_phone = message['from']
        
        # Handle different message types
        message_type = message.get('type')
        
        if message_type == 'text':
            text = message['text']['body']
            handle_text_message(from_phone, text)
        
        elif message_type == 'interactive':
            interactive = message['interactive']
            if interactive['type'] == 'button_reply':
                button_text = interactive['button_reply']['title']
                handle_text_message(from_phone, button_text)
            elif interactive['type'] == 'list_reply':
                list_text = interactive['list_reply']['title']
                handle_text_message(from_phone, list_text)
        
        elif message_type == 'image':
            media_id = message['image']['id']
            handle_payment_screenshot(from_phone, media_id)
        
        return jsonify({'status': 'ok'}), 200
    
    except Exception as e:
        print(f"Webhook error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =================== MESSAGE HANDLER ===================

def handle_text_message(phone, message):
    """Process text messages"""
    user_session = db.get_session(phone)
    step = user_session['step']
    data = user_session['data']
    
    new_step, new_data, response = process_bot_logic(phone, step, data, message)
    
    db.save_session(phone, new_step, new_data)
    
    send_bot_response(phone, new_step, new_data, response)

def process_bot_logic(phone, step, data, message):
    """Main bot logic with improved error handling"""
    
    if not data:
        data = {
            'services': [],
            'total': 0,
            'date': None,
            'time': None,
            'name': None
        }
    
    response = ""
    
    # Universal menu command
    if message and message.lower() in ['menu', 'main menu', 'start', 'restart', 'back to menu']:
        user = db.get_user(phone)
        if user and user.get('name'):
            response = f"👋 Welcome back *{user['name']}*!\n\n"
        else:
            response = f"👋 Welcome to *{Config.SALON_NAME}*!\n\n"
        response += "How can I help you today?"
        return 'menu', response
    
    # Menu
    if step == 'menu' or message.lower() in ['hi', 'hello', 'hey', 'namaste']:
        user = db.get_user(phone)
        if user and user.get('name'):
            response = f"👋 Welcome back *{user['name']}*!\n\n"
        else:
            response = f"👋 Welcome to *{Config.SALON_NAME}*!\n\n"
        
        response += "How can I help you today?"
        step = 'menu'
        data = {}
    
    # New Booking
    elif message in ["📅 New Booking", "Book Now", "New Booking"]:
        user = db.get_user(phone)
        if not user or not user.get('name'):
            response = "Please enter your name:"
            step = 'get_name'
        else:
            data['name'] = user['name']
            response = format_service_list()
            response += "📝 *How to select:*\n"
            response += "Reply with service numbers separated by commas\n"
            response += "Example: 1,3,5\n\n"
            response += "Type *Menu* to go back"
            step = 'select_services'
    
    # Get Name
    elif step == 'get_name':
        name = message.strip()
        if len(name) < 2:
            response = "❌ Please enter a valid name (at least 2 characters):"
            return step, data, response
        
        if name.lower() == 'menu':
            return 'menu', {}, "Returning to main menu..."
        
        data['name'] = name
        db.save_user(phone, name)
        
        response = f"Nice to meet you, *{name}*! 😊\n\n"
        response += format_service_list()
        response += "📝 *How to select:*\n"
        response += "Reply with service numbers separated by commas\n"
        response += "Example: 1,3,5\n\n"
        response += "Type *Menu* to cancel"
        step = 'select_services'
    
    # Select Services - THIS WAS MISSING SECTION
    elif step == 'select_services':
        if message.lower() in ['cancel', 'back']:
            return 'menu', {}, "❌ Booking cancelled. Returning to main menu..."
        
        if not is_valid_service_input(message):
            response = "❌ *Invalid format!*\n\n"
            response += "Please enter service numbers separated by commas.\n"
            response += "Example: 1,3,5\n\n"
            response += format_service_list()
            response += "Or type *Menu* to cancel"
            return step, data, response
        
        try:
            # Parse service numbers
            service_nums = [s.strip() for s in message.replace(' ', '').split(',')]
            data['services'] = service_nums
            
            # Calculate total
            total = sum([Config.SERVICES[s]['price'] for s in service_nums])
            data['total'] = total
            
            # Show selected services summary
            service_names = [Config.SERVICES[s]['name'] for s in service_nums]
            response = "✅ *Selected Services:*\n\n"
            for i, s_num in enumerate(service_nums):
                service = Config.SERVICES[s_num]
                response += f"{i+1}. {service['name']}\n"
                response += f"   💰 ₹{service['price']} | ⏱️ {service['duration']}\n\n"
            
            response += f"*Total Amount:* ₹{total}\n\n"
            
            if total >= Config.ADVANCE_PAYMENT_THRESHOLD:
                advance = int(total * Config.ADVANCE_PERCENTAGE)
                data['advance_required'] = advance
                response += f"💳 *Advance Payment Required:* ₹{advance} (50%)\n\n"
            
            response += "📅 *Select your preferred date:*"
            step = 'select_date'
            
        except Exception as e:
            print(f"Service selection error: {e}")
            response = "❌ Something went wrong. Please try again.\n\n"
            response += "Type *Menu* to start over"
            step = 'menu'
    
    # Select Date
    elif step == 'select_date':
        if message.lower() in ['cancel', 'menu']:
            return 'menu', {}, "❌ Booking cancelled."
        
        # Check if it's a valid date
        dates = get_next_7_days()
        valid_dates = [d['value'] for d in dates]
        
        if message not in valid_dates:
            response = "❌ Invalid date selected.\n\n"
            response += "Please select a date from the list."
            return step, data, response
        
        data['date'] = message
        
        # Get available slots for this date
        available_slots = get_available_slots(message)
        
        if not available_slots:
            response = "😔 Sorry! No slots available on this date.\n\n"
            response += "Please select another date:"
            return step, data, response
        
        # Show available time slots
        response = f"📅 *Date:* {message}\n\n"
        response += "*⏰ Available Time Slots:*\n\n"
        for i, slot in enumerate(available_slots, 1):
            response += f"{i}. {slot}\n"
        
        response += f"\n💡 Reply with slot number (1-{len(available_slots)})\n"
        response += "Or type *Back* to change date"
        
        data['available_slots'] = available_slots
        step = 'select_time'
    
    # Select Time
    elif step == 'select_time':
        if message.lower() == 'back':
            response = "📅 *Select your preferred date:*"
            step = 'select_date'
            return step, data, response
        
        if message.lower() == 'menu':
            return 'menu', {}, "❌ Booking cancelled."
        
        try:
            slot_num = int(message.strip())
            available_slots = data.get('available_slots', [])
            
            if slot_num < 1 or slot_num > len(available_slots):
                response = f"❌ Please enter a valid slot number (1-{len(available_slots)})"
                return step, data, response
            
            selected_time = available_slots[slot_num - 1]
            data['time'] = selected_time
            
            # Show booking summary
            service_names = [Config.SERVICES[s]['name'] for s in data['services']]
            response = "*📋 Booking Summary:*\n\n"
            response += f"👤 *Name:* {data['name']}\n"
            response += f"📅 *Date:* {data['date']}\n"
            response += f"⏰ *Time:* {data['time']}\n\n"
            response += "*Services:*\n"
            for s in data['services']:
                response += f"• {Config.SERVICES[s]['name']} - ₹{Config.SERVICES[s]['price']}\n"
            response += f"\n💰 *Total:* ₹{data['total']}\n\n"
            
            # Check if advance payment required
            if data['total'] >= Config.ADVANCE_PAYMENT_THRESHOLD:
                advance = data.get('advance_required', 0)
                response += f"💳 *Advance Required:* ₹{advance}\n\n"
                response += "Click *Proceed to Payment* to continue"
                step = 'confirm_with_payment'
            else:
                response += "✅ No advance payment required!\n\n"
                response += "Click *Confirm Now* to book your appointment"
                step = 'confirm_without_payment'
            
        except ValueError:
            response = "❌ Please enter a valid number"
            return step, data, response
        except Exception as e:
            print(f"Time selection error: {e}")
            response = "❌ Something went wrong. Please try again."
            return step, data, response
    
    # Confirm without payment
    elif step == 'confirm_without_payment':
        if message in ["✅ Confirm Now", "Confirm", "Yes"]:
            # Create booking
            booking_id = db.save_booking(
                phone=phone,
                name=data['name'],
                services=data['services'],
                date=data['date'],
                time=data['time'],
                total=data['total'],
                advance_required=0,
                status='confirmed'
            )
            
            service_names = [Config.SERVICES[s]['name'] for s in data['services']]
            response = "🎉 *Booking Confirmed!*\n\n"
            response += f"*Booking ID:* #{booking_id}\n"
            response += f"*Name:* {data['name']}\n"
            response += f"*Date:* {data['date']}\n"
            response += f"*Time:* {data['time']}\n"
            response += f"*Services:* {', '.join(service_names)}\n"
            response += f"*Total:* ₹{data['total']}\n\n"
            response += f"✨ See you at *{Config.SALON_NAME}*!\n\n"
            response += f"📍 {Config.SALON_ADDRESS}\n"
            response += f"📞 {Config.SALON_PHONE}\n\n"
            response += "Type *Menu* for more options"
            
            step = 'menu'
            data = {}
        
        elif message in ["❌ Cancel", "Cancel"]:
            response = "❌ Booking cancelled.\n\nType *Menu* to start over"
            step = 'menu'
            data = {}
        else:
            response = "Please click *Confirm Now* or *Cancel*"
            return step, data, response
    
    # Confirm with payment
    elif step == 'confirm_with_payment':
        if message in ["💳 Proceed to Payment", "Proceed", "Pay"]:
            response = "📱 *Payment Information*\n\n"
            response += f"*Amount to Pay:* ₹{data.get('advance_required', 0)}\n"
            response += f"*UPI ID:* {Config.UPI_ID}\n\n"
            response += "I'll send you the QR code in the next message. 👇"
            step = 'show_payment'
        
        elif message in ["❌ Cancel", "Cancel"]:
            response = "❌ Booking cancelled.\n\nType *Menu* to start over"
            step = 'menu'
            data = {}
        else:
            response = "Please click *Proceed to Payment* or *Cancel*"
            return step, data, response
    
    # Show payment and wait for screenshot
    elif step == 'show_payment':
        if message in ["✅ I Have Paid", "Paid", "Done"]:
            response = "📸 *Please upload your payment screenshot*\n\n"
            response += "Take a screenshot of your payment confirmation and send it here.\n\n"
            response += "We'll verify and confirm your booking within *1 hour*. ⏰"
            step = 'waiting_payment_screenshot'
        
        elif message in ["🔙 Back", "Back"]:
            service_names = [Config.SERVICES[s]['name'] for s in data['services']]
            response = "*📋 Booking Summary:*\n\n"
            response += f"👤 *Name:* {data['name']}\n"
            response += f"📅 *Date:* {data['date']}\n"
            response += f"⏰ *Time:* {data['time']}\n\n"
            response += f"💰 *Total:* ₹{data['total']}\n"
            response += f"💳 *Advance:* ₹{data.get('advance_required', 0)}\n\n"
            response += "Click *Proceed to Payment* to continue"
            step = 'confirm_with_payment'
        else:
            response = "Please click *I Have Paid* after completing payment,\nor click *Back* to review your booking."
            return step, data, response
    
    # Waiting for payment screenshot
    elif step == 'waiting_payment_screenshot':
        if message.lower() == 'menu':
            response = "⚠️ *Booking In Progress*\n\n"
            response += "Please upload your payment screenshot to complete the booking.\n\n"
            response += "Type *Cancel* if you want to cancel this booking."
            return step, data, response
        
        elif message.lower() == 'cancel':
            response = "❌ Booking cancelled.\n\nType *Menu* to start over"
            step = 'menu'
            data = {}
        else:
            response = "📸 Please send the *payment screenshot* as an image.\n\n"
            response += "Or type *Cancel* to cancel this booking."
            return step, data, response
    
    # My Bookings
    elif message in ["📋 My Bookings", "My Bookings", "Bookings"]:
        bookings = db.get_bookings(phone=phone)
        
        if bookings:
            response = f"*📋 Your Bookings ({len(bookings)}):*\n\n"
            
            for i, booking in enumerate(bookings[:5], 1):  # Show first 5
                services = json.loads(booking['services'])
                service_names = [Config.SERVICES[s]['name'] for s in services]
                
                response += f"*#{booking['id']}* - {booking['date']} at {booking['time']}\n"
                response += f"💇 {', '.join(service_names)}\n"
                response += f"💰 ₹{booking['total']}\n"
                response += f"Status: {booking['status'].replace('_', ' ').title()}\n\n"
            
            if len(bookings) > 5:
                response += f"...and {len(bookings)-5} more\n\n"
        else:
            response = "📭 You don't have any bookings yet.\n\n"
            response += "Book your first appointment now!"
        
        response += "\nType *Menu* to go back"
        step = 'menu'
    
    # Contact
    elif message in ["📞 Contact Us", "Contact", "Contact Us"]:
        response = f"*📞 Contact {Config.SALON_NAME}*\n\n"
        response += f"📍 *Address:*\n{Config.SALON_ADDRESS}\n\n"
        response += f"📱 *Phone:*\n{Config.SALON_PHONE}\n\n"
        response += f"💳 *UPI ID:*\n{Config.UPI_ID}\n\n"
        response += "*🕐 Working Hours:*\n"
        response += "Monday - Sunday\n"
        response += "10:00 AM - 8:00 PM\n\n"
        response += "Type *Menu* to go back"
        step = 'menu'
    
    # Default
    else:
        response = "❓ I didn't understand that.\n\n"
        response += "Type *Menu* to see all options\n"
        response += "or choose from:\n"
        response += "• *New Booking* - Book appointment\n"
        response += "• *My Bookings* - View bookings\n"
        response += "• *Contact Us* - Get contact info"
    
    return step, data, response

def handle_payment_screenshot(phone, media_id):
    """Handle payment screenshot upload"""
    user_session = db.get_session(phone)
    step = user_session['step']
    data = user_session['data']
    
    if step == 'waiting_payment_screenshot':
        filename = f"payment_{phone}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        save_path = os.path.join(Config.UPLOAD_FOLDER, filename)
        
        downloaded = whatsapp.download_media(media_id, save_path)
        
        if downloaded:
            booking_id = db.save_booking(
                phone=phone,
                name=data['name'],
                services=data['services'],
                date=data['date'],
                time=data['time'],
                total=data['total'],
                advance_required=data.get('advance_required', 0),
                status='payment_pending'
            )
            
            db.update_booking(booking_id, payment_screenshot=filename)
            
            response = "✅ *Payment Screenshot Received!*\n\n"
            response += f"*Booking ID:* #{booking_id}\n\n"
            response += "🔍 *Under Review*\n"
            response += "Our team will verify your payment and confirm within *1 hour*.\n\n"
            response += "You'll receive a confirmation message once approved. 🎉\n\n"
            response += "Type *Menu* for more options"
            
            whatsapp.send_message(phone, response)
            
            db.save_session(phone, 'menu', {})
        else:
            whatsapp.send_message(phone, "❌ Failed to receive image. Please try uploading again.")

def send_bot_response(phone, step, data, text_response):
    """Send appropriate response based on step"""
    
    if step == 'menu':
        buttons = ["📅 New Booking", "📋 My Bookings", "📞 Contact Us"]
        whatsapp.send_interactive_buttons(phone, text_response, buttons)
    
    elif step == 'select_date':
        dates = get_next_7_days()
        sections = [{
            "title": "Available Dates",
            "rows": [{"id": d['value'], "title": d['label']} for d in dates]
        }]
        whatsapp.send_interactive_list(phone, text_response, "📅 Choose Date", sections)
    
    elif step == 'confirm_without_payment':
        whatsapp.send_interactive_buttons(phone, text_response, ["✅ Confirm Now", "❌ Cancel"])
    
    elif step == 'confirm_with_payment':
        whatsapp.send_interactive_buttons(phone, text_response, ["💳 Proceed to Payment", "❌ Cancel"])
    
    elif step == 'show_payment':
        advance = data.get('advance_required', 0)
        caption = f"*💳 Payment Required*\n\n"
        caption += f"*Amount:* ₹{advance}\n"
        caption += f"*UPI ID:* {Config.UPI_ID}\n\n"
        caption += "📱 Scan the QR code above to pay.\n\n"
        caption += "After payment, click *I Have Paid* button"
        
        whatsapp.send_image(phone, Config.QR_CODE_PATH, caption)
        whatsapp.send_interactive_buttons(phone, "Have you completed the payment?", ["✅ I Have Paid", "🔙 Back"])
    
    else:
        if text_response:
            whatsapp.send_message(phone, text_response)

# =================== ADMIN PANEL ===================

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        if password == Config.ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Invalid password!', 'error')
    
    return render_template('admin.html', page='login')

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    bookings = db.get_bookings()
    
    stats = {
        'total': len(bookings),
        'confirmed': len([b for b in bookings if b['status'] == 'confirmed']),
        'pending': len([b for b in bookings if b['status'] == 'payment_pending']),
        'cancelled': len([b for b in bookings if b['status'] in ['cancelled', 'rejected']])
    }
    
    return render_template('admin.html', 
                         page='dashboard', 
                         bookings=bookings, 
                         stats=stats,
                         services=Config.SERVICES)

@app.route('/admin/booking/<int:booking_id>/approve', methods=['POST'])
def approve_booking(booking_id):
    if 'admin_logged_in' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    booking = db.get_booking(booking_id)
    if booking:
        db.update_booking(booking_id, status='confirmed')
        
        services = json.loads(booking['services'])
        service_names = [Config.SERVICES[s]['name'] for s in services]
        
        message = "🎉 *Payment Verified - Booking Confirmed!*\n\n"
        message += f"*Booking ID:* #{booking_id}\n"
        message += f"*Name:* {booking['name']}\n"
        message += f"*Date:* {booking['date']}\n"
        message += f"*Time:* {booking['time']}\n"
        message += f"*Services:* {', '.join(service_names)}\n"
        message += f"*Total:* ₹{booking['total']}\n\n"
        message += f"✨ See you at *{Config.SALON_NAME}*!\n\n"
        message += f"📍 {Config.SALON_ADDRESS}\n"
        message += f"📞 {Config.SALON_PHONE}"
        
        whatsapp.send_message(booking['phone'], message)
        
        flash(f'Booking #{booking_id} approved and customer notified!', 'success')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/booking/<int:booking_id>/reject', methods=['POST'])
def reject_booking(booking_id):
    if 'admin_logged_in' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    notes = request.form.get('notes', 'Payment verification failed')
    booking = db.get_booking(booking_id)
    
    if booking:
        db.update_booking(booking_id, status='rejected', admin_notes=notes)
        
        message = "❌ *Booking Payment Rejected*\n\n"
        message += f"*Booking ID:* #{booking_id}\n"
        message += f"*Reason:* {notes}\n\n"
        message += "Please contact us for clarification:\n"
        message += f"📞 {Config.SALON_PHONE}\n\n"
        message += "You can rebook by typing *New Booking*"
        
        whatsapp.send_message(booking['phone'], message)
        
        flash(f'Booking #{booking_id} rejected!', 'warning')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

# =================== EXPORT & PRINT FEATURES ===================

@app.route('/admin/export/excel')
def export_excel():
    """Export all bookings to Excel"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file
    
    # Get all bookings
    bookings = db.get_bookings()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Salon Bookings"
    
    # Headers
    headers = ['Booking ID', 'Date', 'Time', 'Customer Name', 'Phone', 'Services', 
               'Total Amount', 'Advance', 'Status', 'Payment Screenshot', 'Created At', 'Admin Notes']
    
    # Style headers
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, booking in enumerate(bookings, 2):
        services = json.loads(booking['services'])
        service_names = ', '.join([Config.SERVICES[s]['name'] for s in services])
        
        ws.cell(row=row_idx, column=1, value=booking['id'])
        ws.cell(row=row_idx, column=2, value=booking['date'])
        ws.cell(row=row_idx, column=3, value=booking['time'])
        ws.cell(row=row_idx, column=4, value=booking['name'])
        ws.cell(row=row_idx, column=5, value=booking['phone'])
        ws.cell(row=row_idx, column=6, value=service_names)
        ws.cell(row=row_idx, column=7, value=booking['total'])
        ws.cell(row=row_idx, column=8, value=booking['advance_required'] if booking['advance_required'] else 0)
        ws.cell(row=row_idx, column=9, value=booking['status'])
        ws.cell(row=row_idx, column=10, value=booking['payment_screenshot'] if booking['payment_screenshot'] else 'N/A')
        ws.cell(row=row_idx, column=11, value=booking['created_at'])
        ws.cell(row=row_idx, column=12, value=booking['admin_notes'] if booking['admin_notes'] else '')
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"salon_bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/export/pdf')
def export_pdf():
    """Export bookings to PDF"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from flask import send_file
    
    # Get bookings
    bookings = db.get_bookings()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"<b>{Config.SALON_NAME}</b><br/>Booking Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Stats summary
    stats_data = [
        ['Total Bookings', 'Confirmed', 'Pending Payment', 'Cancelled'],
        [
            str(len(bookings)),
            str(len([b for b in bookings if b['status'] == 'confirmed'])),
            str(len([b for b in bookings if b['status'] == 'payment_pending'])),
            str(len([b for b in bookings if b['status'] in ['cancelled', 'rejected']]))
        ]
    ]
    
    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Bookings table
    data = [['ID', 'Date', 'Time', 'Customer', 'Phone', 'Services', 'Amount', 'Status']]
    
    for booking in bookings[:50]:  # Limit to 50 for PDF
        services = json.loads(booking['services'])
        service_names = ', '.join([Config.SERVICES[s]['name'] for s in services])
        
        data.append([
            str(booking['id']),
            booking['date'],
            booking['time'],
            booking['name'],
            booking['phone'][-4:],  # Last 4 digits
            service_names[:30] + '...' if len(service_names) > 30 else service_names,
            f"₹{booking['total']}",
            booking['status']
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 0.3*inch))
    footer_text = f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"salon_bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/booking/<int:booking_id>/print')
def print_booking(booking_id):
    """Generate printable invoice for a booking"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from io import BytesIO
    from flask import send_file
    
    booking = db.get_booking(booking_id)
    
    if not booking:
        flash('Booking not found!', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Header style
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading1'],
        fontSize=26,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=12,
        alignment=1
    )
    
    # Salon name
    header = Paragraph(f"<b>{Config.SALON_NAME}</b>", header_style)
    elements.append(header)
    
    # Salon details
    salon_info = Paragraph(
        f"{Config.SALON_ADDRESS}<br/>{Config.SALON_PHONE}",
        styles['Normal']
    )
    elements.append(salon_info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Invoice title
    invoice_title = Paragraph(
        f"<b>BOOKING INVOICE</b><br/>Booking ID: #{booking_id}",
        ParagraphStyle('InvoiceTitle', parent=styles['Heading2'], alignment=1, fontSize=18, spaceAfter=20)
    )
    elements.append(invoice_title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Customer details
    customer_data = [
        ['Customer Name:', booking['name']],
        ['Phone:', booking['phone']],
        ['Date:', booking['date']],
        ['Time:', booking['time']],
        ['Booking Status:', booking['status'].replace('_', ' ').title()]
    ]
    
    customer_table = Table(customer_data, colWidths=[2*inch, 4*inch])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Services table
    services = json.loads(booking['services'])
    service_data = [['Service', 'Duration', 'Price']]
    
    for s_id in services:
        service = Config.SERVICES[s_id]
        service_data.append([
            service['name'],
            service['duration'],
            f"₹{service['price']}"
        ])
    
    # Add total row
    service_data.append(['', 'TOTAL:', f"₹{booking['total']}"])
    
    service_table = Table(service_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    service_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('GRID', (0, 0), (-1, -2), 1, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
    ]))
    elements.append(service_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Payment info
    if booking['advance_required'] > 0:
        payment_info = Paragraph(
            f"<b>Advance Paid:</b> ₹{booking['advance_required']}<br/>"
            f"<b>Balance Due:</b> ₹{booking['total'] - booking['advance_required']}",
            styles['Normal']
        )
        elements.append(payment_info)
        elements.append(Spacer(1, 0.2*inch))
    
    # Footer
    footer = Paragraph(
        f"<i>Thank you for choosing {Config.SALON_NAME}!<br/>"
        f"We look forward to serving you.<br/>"
        f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}</i>",
        ParagraphStyle('Footer', parent=styles['Normal'], alignment=1, fontSize=9)
    )
    elements.append(Spacer(1, 0.5*inch))
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"booking_{booking_id}_invoice.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/reports')
def admin_reports():
    """Advanced reporting page"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status')
    
    bookings = db.get_bookings()
    
    # Apply filters
    if start_date:
        bookings = [b for b in bookings if b['date'] >= start_date]
    if end_date:
        bookings = [b for b in bookings if b['date'] <= end_date]
    if status_filter:
        bookings = [b for b in bookings if b['status'] == status_filter]
    
    # Calculate statistics
    total_revenue = sum([b['total'] for b in bookings if b['status'] == 'confirmed'])
    total_advance = sum([b['advance_required'] for b in bookings if b['advance_required']])
    
    # Service popularity
    service_count = {}
    for booking in bookings:
        services = json.loads(booking['services'])
        for s_id in services:
            service_name = Config.SERVICES[s_id]['name']
            service_count[service_name] = service_count.get(service_name, 0) + 1
    
    # Most popular services
    popular_services = sorted(service_count.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Date-wise bookings
    date_bookings = {}
    for booking in bookings:
        date = booking['date']
        date_bookings[date] = date_bookings.get(date, 0) + 1
    
    stats = {
        'total_bookings': len(bookings),
        'confirmed': len([b for b in bookings if b['status'] == 'confirmed']),
        'pending': len([b for b in bookings if b['status'] == 'payment_pending']),
        'cancelled': len([b for b in bookings if b['status'] in ['cancelled', 'rejected']]),
        'total_revenue': total_revenue,
        'total_advance': total_advance,
        'popular_services': popular_services,
        'date_bookings': date_bookings
    }
    
    return render_template('reports.html', 
                         stats=stats, 
                         bookings=bookings,
                         start_date=start_date,
                         end_date=end_date,
                         status_filter=status_filter)

@app.route('/admin/backup')
def backup_database():
    """Download database backup"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from flask import send_file
    
    db_path = db.db_name
    filename = f"salon_db_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    
    return send_file(
        db_path,
        mimetype='application/octet-stream',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin/booking/<int:booking_id>/delete', methods=['POST'])
def delete_booking(booking_id):
    """Delete a booking"""
    if 'admin_logged_in' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    booking = db.get_booking(booking_id)
    if booking:
        db.delete_booking(booking_id)
        flash(f'Booking #{booking_id} deleted successfully!', 'success')
    else:
        flash('Booking not found!', 'error')
    
    return redirect(url_for('admin_dashboard'))

# =================== HEALTH CHECK ===================

@app.route('/')
def home():
    return jsonify({
        'status': 'active',
        'app': Config.SALON_NAME,
        'webhook': '/webhook',
        'admin': '/admin'
    })

@app.route('/health')
def health():
    return jsonify({'status': 'healthy'}), 200

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT",5000))
    app.run(host='0.0.0.0', port=port)
